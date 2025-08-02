from csv import excel

import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side

PATH = "/Users/sergej/Documents/ExcelTest/"

def check_percent(data1, data2):
    result = 0
    try:
        result = (((data1 / data2 )-1) * 100)
    except ZeroDivisionError:
        # print("Rais Zero Division")
        return 0
    except ValueError:
        # print("Rais Value Error")
        return 0
    return result

def convert_to_dict(pd_array: pd.DataFrame) -> dict:
    result_dict = {}
    for index, row in pd_array.iterrows():
        result_dict[index] = [pd_array.iloc[index, 0], pd_array.fillna(0).iloc[index, 7]]

    return result_dict
def add_product(list, name, value_doc1, value_doc2):
    diff_value = value_doc2 - value_doc1
    list.append([name, value_doc2, value_doc1, diff_value, check_percent(value_doc2, value_doc1)])

def equal_products(doc1, doc2):
    data1_dict = convert_to_dict(doc1)
    data2_dict = convert_to_dict(doc2)
    sheet_name = ''

    for index1, row1 in list(data1_dict.items()):
        for index2, row2 in list(data2_dict.items()):
            if row1[0] in exclude_list:
                break
            elif row1[0] == row2[0]:
                # print(f"Row name: {row1[0]} = Value: {row1[1]} \t Row name: {row2[0]} = Value: {row2[1]}")
                data2_dict.pop(index2)
                if row1[0] == "Цех №1":
                    sheet_name = "Цех №1"
                    print(f"sheet name: {sheet_name}")
                elif row1[0] == "Цех №2":
                    sheet_name = "Цех №2"
                    print(f"sheet name: {sheet_name}")
                elif row1[0] == "Цех №3":
                    sheet_name = "Цех №3"
                    print(f"sheet name: {sheet_name}")
                elif row1[0] == "Аутсорсинг":
                    sheet_name = "Аутсорсинг"
                    print(f"sheet name: {sheet_name}")

                if sheet_name == "Цех №1" and row1[1] != row2[1]:
                    try:
                        add_product(mf1_list, row1[0], row1[1], row2[1])
                        break
                    except TypeError as ex:
                        print(ex)
                elif  sheet_name == "Цех №2" and row1[1] != row2[1]:
                    try:
                        add_product(mf2_list, row1[0], row1[1], row2[1])
                        # mf2_list.append([row1[0], row2[1], row1[1], row2[1] - row1[1], check_percent(row2[1], row1[1])])
                        break
                    except TypeError as ex:
                        print(ex)
                elif sheet_name == "Цех №3" and row1[1] != row2[1]:
                    try:
                        add_product(mf3_list, row1[0], row1[1], row2[1])
                        # mf3_list.append([row1[0], row2[1], row1[1], row2[1] - row1[1], check_percent(row2[1], row1[1])])
                        break
                    except TypeError as ex:
                        print(ex)
                elif sheet_name == "Итого":
                    break
                # else:
                #     try:
                #         outsource_list.append([row1[0], row2[1], row1[1], row2[1] - row1[1], check_percent(row2[1], row1[1])])
                #         break
                #     except TypeError as ex:
                #         print(ex)
            continue

def not_equal_products(data1, data2):
    result_dict = {}
    for index1, row in data1.iterrows():
        for index2, col in data2.iterrows():
            if data1.iloc[index1,0] != data2.iloc[index2,0]:
                print(data1.iloc[index1,0])

def styler(row, font_size: int, font_weight: bool, color, indent: int = 0,):
    print(f"Row: {row} = {row[0]}")
    row[0].alignment = Alignment(vertical='center', indent=indent)
    for cell in row:
        cell.border = border
        cell.font = Font(size=font_size, bold=font_weight)
        # cell.alignment = Alignment(vertical='center', indent=indent)
        cell.fill = PatternFill("solid", fgColor=color, fill_type='solid')

doc1_name = "Остаток ГП и ПФ на 01.07.24.xlsx"
doc2_name = "Остаток ГП и ПФ на 01.07.25.xlsx"
sheet_name = ''
exclude_list = ["Оценка избыточности запасов ПФ и ГП", "Параметры:", "Отбор:", "Цех основной продукции"]
data_columns = ["Номенклатура", doc2_name[0:-5], doc1_name[0:-5], "Разница", "Процент"]
dept_list = ["Цех №1" , "Цех №2", "Цех №3"]

mf_goodies_lvl2 = ('Полуфабрикаты', 'Готовая Продукция', 'Удаленная и не используемая продукция' )

mf_goodies_lvl3 = ['Полуфабрикаты т.м. СТМ', 'Полуфабрикаты т.м. Пиканта', 'Полуфабрикаты т.м. Ресторация Обломов', 'Полуфабрикаты т.м. Угощение Славянки',
                     'Полуфабрикаты т.м. BALKANIKA', 'Полуфабрикаты т.м. Бабулины Рецепты', 'Заморозка , Концентраты, Маринады', 'Полуфабрикаты Цех№3',
                     'Концентраты в бочках 200 л.', 'Готовая продукция Цех №1', 'Готовая продукция_АКК', 'Готовая продукция Цех №2', 'Полуфабрикаты',
                     'Готовая продукция (пюре) по 6 шт. 530 г.', 'Сладкая группа', 'Готовая Продукция т.м. _СТМ', 'Полуфабрикаты Цех№3', 'Готовая продукция Цех №1',
                     'Готовая продукция Цех №3', 'Готовая продукция Цех №2', 'Полуфабрикаты Цех№2', 'Готовая продукция (пюре) по 6 шт. 530 г.',
                     'ЗАМОРОЖЕННАЯ ПРОДУКЦИЯ_АКК (под продажу)','Товар Аутсорсинг']

mf_goodies_lvl4 = ['Овощная консервация СТМ', 'Икра 0,7 ТВ СТМ', 'Икра 0,5 СТМ', 'Икра Ж/Б №9', 'Икра 0,35 СТМ', 'Икра 0,35 ТВ СТМ', 'Икра 0,39 СТМ в банке Банка Славянка',
                   'Икра 0,45 СТМ банке КБ 202 Баб рец', 'Итальянские томаты', 'Овощная Консервация 0,42 ТВ Пиканта', 'Овощная консервация 0,7 ТВ Пиканта',
                   'Маринады', 'Овощная консервация 0,5 ТВ Пиканта', 'Овощная Консервация 0,35 ТВ Пиканта', 'Концентраты', 'Овощная Консервация 0,45 ТВ Пиканта',
                   'Овощная Консервация Пиканта de Luxe', 'Овощная консервация ж/б', 'Овощная Консервация (Икра 0,5 ТВ)', 'Овощная консервация 0,30 ТВ',
                   'Овощная консервация 0,4 (Икра)', 'Овощная консервация 0,42 ТВ', 'Овощная консервация 0,25 ТВ (спреды)', 'Овощная консервация 0,25 ТВ',
                   'Овощная консервация Уг.Слав.', 'Икра BALKANIKA', 'Лето BALKANIKA', 'Лето 0,45 Бабулины Рецепты', 'Лето 0,5 Бабулины Рецепты',
                   'Быстрозамороженная продукция', 'Полуфабрикаты маринады в бочках', 'Хорека', 'Готовая Продукция_СТМ Цех №1', 'Готовая Продукция ЭКСПОРТ Цех №1',
                   'Готовая Продукция т.м. Пиканта', 'Готовая Продукция т.м. Угощение Славянки', 'Готовая Продукция т.м."Ресторация Обломов"', 'Готовая Продукция т.м. Бабулины Рецепты',
                   'Готовая Продукция т.м. BALKANIKA', 'Готовая продукция т.м. Пиканта de Luxe', 'Готовая Продукция Ж/Б т.м. Пиканта', 'СТМ_АКК', 'Готовая продукция_Хорека',
                   'ЭКСПОРТ СЛАДКАЯ ГРУППА', 'Маринады 0,45 ТВ СТМ', 'Готовая продукция "КРУГЛЫЙ ГОД"', 'Итальянские томаты_СТМ', 'Итальянские томаты_Пиканта', 'Готовая Продукция_СТМ Цех №1',
                   'Готовая Продукция ЭКСПОРТ Цех №1', 'Готовая Продукция т.м. Пиканта', 'Готовая продукция_Хорека', 'Готовая продукция т.м. Пиканта_Дойпак', 'Готовая продукция Baresto (Сиропы)',
                   'Готовая продукция_СТМ Цех №2', 'Готовая продукция Пиканта(сиропы)', 'Готовая продукция Пиканта(пюре)', 'Полуфабрикаты_Baresto', 'Полуфабрикаты_Пиканта', 'Полуфабрикаты_СТМ',
                   'Сырье замороженное']

mf_goodies_lvl5 = ['Лето 0,5 СТМ', 'Зима 0,35 СТМ', 'Лето 0,7 СТМ', 'Зима 0,5 СТМ', 'Лето 0,35 СТМ', 'Зима 0,45 СТМ Банка (450-Р12.61.3-ВОКЭ-18-012)', 'Лето 0,45 СТМ Банка (450-Р12.61.3-ВОКЭ-18-012)',
                   'Лето 0,4 СТМ', 'Ж/б  СТМ', 'Зима 0,39 СТМ в банке Славянка', 'Зима 0,45 СТМ в банке КБ 202 Баб рец', 'Зима 0,7 СТМ', 'Лето 0,25 СТМ', 'Лето 0,39 СТМ в банке Банка Славянка',
                   'Лето 0,45 СТМ в банке КБ 202 Баб рец', 'Овощная консервация 0,42 ТВ Пиканта (бобовая группа)', 'Овощная консервация 0,42 ТВ Пиканта (закусочная группа)', 'Овощная консервация 0,42 ТВ Пиканта (нарезная группа)',
                   'Овощная консервация 0,42 ТВ Пиканта (группа лечо)', 'Овощная консервация 0,42 ТВ Пиканта (икорная группа)', 'Маринады 0,7 ТВ', 'Маринады 0,42 ТВ', 'Маринады 0,45 ТВ', 'Итальянские томаты', 'Зима Пиканта',
                   'Лето Пиканта', 'Зима Пиканта', 'Ж/б 8', '0,7 ТВ Угощ Слав', 'Ж/б 12', 'Лето Обломов', 'Зима Обломов', 'Икра Обломов', 'Итальянские томаты', 'Овощная консервация', 'Готовая Продукция_Global Village по 6 шт',
                   "Готовая Продукция_Монетка (O'GREEN)", 'Готовая Продукция_SPAR', 'Готовая Продукция_ВЕРНЫЙ по 6 шт.', 'Готовая Продукция_Самокат', 'Готовая Продукция_Магнит', 'Готовая Продукция_Праздник урожая',
                   'Готовая Продукция_Умный выбор', 'Готовая продукция_Vitalgreens', 'Готовая Продукция_Лента', 'Готовая Продукция_Амбар Победа/Большая грядка', 'Готовая Продукция_АШАН', 'Готовая продукция_ВкусВилл',
                   'Готовая продукция_ГЛОБУС', 'Готовая продукция_Из Лавки', 'Готовая Продукция_ПРОСТО АЗБУКА', 'Готовая Продукция_Маркет Перекресток по 6 шт.', 'Готовая Продукция_ВЕЛАДА_Мария Ра', 'Готовая продукция_GUSTERRO',
                   'Готовая Продукция_Домашние Заготовки', 'Готовая Продукция_Кировский', 'Готовая Продукция_Овощная Семейка', 'Готовая Продукция_Сладкая жизнь', 'Готовая Продукция Asim (Израиль)', 'Готовая Продукция_SPAR Скиф Трейд ООО (Казахстан)',
                   'Готовая Продукция Фея Маури (Киргизия)', 'Готовая Продукция Lackmann Fleisch (Германия)', 'Готовая Продукция ФЛ Центр (Беларусь)', 'Готовая Продукция ФЛ Центр (Армения)', 'Готовая Продукция BALTON TRADING (ASIA) (Узбекистан)',
                   'Готовая Продукция ПВ Азия (Казахстан)', 'Готовая Продукция DalaProgect (Казахстан)', 'Готовая Продукция SLCO GmbH & Co. KG (Германия)', 'Готовая Продукция Араз ООО (Азербайджан)', 'Готовая Продукция ПВ-Запад (Беларусь) т.м. Пиканта',
                   'Готовая Продукция ФЛ-Центр (Киргизия)', 'Готовая Продукция Экофорт (Беларусь)', 'Готовая Продукция ТМ Угощение Славянки Ж/Б по 6 шт', 'Готовая Продукция ТМ Угощение Славянки СТЕКЛО по 6 шт', 'Готовая Продукция ТМ Угощение Славянки Ж/Б по 12 шт',
                   'Готовая Продукция ТМ Угощение Славянки СТЕКЛО по 12 шт', 'Готовая Продукция т.м. Пиканта по 6 шт. 350-440 гр.', 'Готовая Продукция т.м. Пиканта по 6 шт. 520-550 гр', 'Готовая Продукция т.м. Пиканта_700 гр по 6 шт.',
                   'Готовая продукция т.м. Пиканта_Итальянские томаты', 'Готовая Продукция т.м."Ресторация Обломов" по 6 шт. 250 мл', 'Готовая Продукция т.м."Ресторация Обломов" по 6 шт. 300 мл', 'Готовая Продукция т.м."Ресторация Обломов" по 6 шт. 400 мл',
                   'Готовая Продукция т.м."Ресторация Обломов" по 6 шт. 300 мл Банка BALKANIKA', 'Готовая Продукция т.м. Бабулины Рецепты по 6 шт.', 'Готовая продукция т.м. Пиканта de Luxe по 6 шт', 'ГРЯДКА УДАЧИ и КД', 'Итальянские томаты', 'Бочки',
                   'Итальянские томаты', 'Готовая продукция_Хорека_Итальянские томаты', 'Готовая Продукция Экофорт (Беларусь)', 'Готовая продукция т.м. Пиканта_Итальянские томаты', 'Готовая продукция СТМ_Итальянские томаты', 'Готовая продукция Экофорт (Беларусь)',
                   'Готовая Продукция Араз ООО (Азербайджан) т.м. Пиканта Дойпак', 'Готовая продукция BALTON TRADING (ASIA) (Узбекистан)', 'Готовая Продукция ФЛ Центр (Армения) т.м. Пиканта (Дойпак)', 'Готовая Продукция ФЛ Центр (Казахстан) т.м. Пиканта (Дойпак)',
                   'Готовая Продукция ПВ-Запад (Беларусь) т.м. Пиканта (Дойпак)', 'Готовая Продукция_Дамдет ТОО (КАЗАХСТАН)', 'Готовая продукция ГМС ПРО, ООО (Беларусь)  т.м. "Baresto" 250 мл', 'Готовая продукция Фея Маури (Киргизия)', 'Дойпак по 16 шт.',
                   'Дойпак по 10 шт.', 'Дойпак по 12 шт.', 'Готовая продукция т.м. "Baresto" 250 мл', 'Готовая продукция т.м. "Baresto" 1л', 'Готовая продукция т.м. "Baresto" 1л по 12шт.', 'Готовая продукция т.м. "Baresto" 1л по 12шт.', 'Готовая продукция_ВкусВилл',
                   'Готовая продукция Global Village', 'Готовая продукция АШАН', 'Готовая продукция_Лента 1 л', 'Готовая продукция_Лента 250 мл', 'Готовая продукция_ОКЕЙ 1 литр', 'Готовая продукция_ТМ «DELIISE»', 'Готовая продукция_ТМ «SPAR» 1000 мл',
                   'Готовая продукция_ТМ «SPAR» 250 мл', 'Готовая продукция_ТМ PREMIERE of TASTE 250 мл', 'Готовая продукция_ТМ Монетка 250 мл', 'Готовая продукция_ТМ. «МКухня» 1л.', 'Готовая продукция т.м. "Пиканта" 250 мл', 'Готовая продукция т.м. "Пиканта" 250 мл (наборы)',
                   'Готовая продукция Пиканта (пюре) по 6 шт', 'Сиропы_Baresto 0,25л. ТВ', 'Сиропы_Baresto 1 л.', 'Сиропы_Пиканта', 'Десерты_Пиканта', 'Десерты_АШАН', 'Сироп_PREMIERE of TASTE 0,25л.', 'Сиропы_BONVIDA 1л.', 'Сиропы_CENSA 1л.', 'Сиропы_DELIISE 0,25', 'Сиропы_PREMIUM CLUB 0,25л.',
                   'Сиропы_SPAR 0,25л.', 'Сиропы_АШАН 0,25л.', 'Сиропы_Магнит 1л.', 'Сиропы_МаркетПерекресток 0,25л.', 'Сиропы_Монетка 0,25 л.', 'Сиропы_ОКЕЙ 1л.']

mf1_list = []
mf2_list = []
mf3_list = []
outsource_list = []



# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    doc1 = pd.read_excel(PATH + doc1_name)
    doc2 = pd.read_excel(PATH + doc2_name)

    changes = doc2.fillna(0).iloc[15:,7] - doc1.fillna(0).iloc[15:,7]
    changes_pct = check_percent(doc1.iloc[15:,7] , doc2.iloc[15:,7])

    equal_products(doc1, doc2)

    pd_mf1 = pd.DataFrame(mf1_list, columns=data_columns)
    pd_mf2 = pd.DataFrame(mf2_list, columns=data_columns)
    pd_mf3 = pd.DataFrame(mf3_list, columns=data_columns)
    pd_outsource = pd.DataFrame(outsource_list, columns=data_columns)

    with pd.ExcelWriter(PATH + 'TestChanges2.xlsx') as writer:
        pd_mf1.to_excel(writer, index=False, sheet_name="Цех 1", float_format='%.2f')
        pd_mf2.to_excel(writer, index=False, sheet_name="Цех 2", float_format='%.2f')
        pd_mf3.to_excel(writer, index=False, sheet_name="Цех 3", float_format='%.2f')

    wb = openpyxl.load_workbook(PATH + 'TestChanges2.xlsx')
    ws = wb.active


    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = 'A2'
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.column_dimensions['A'].width = 150
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        thin_side = Side(border_style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row in ws.iter_rows():
            if row[0].value == "Номенклатура":
                styler(row, 20,True,'FBE559')
            if row[0].value in dept_list:
                styler(row, 18,True,'FBF2D8')
                # for cell in row:
                #     cell.font = Font(size=18, bold=True)
                #     cell.alignment = Alignment(vertical='center')
                #     cell.fill = PatternFill("solid", fgColor="F8F2D8", fill_type='solid')
            elif row[0].value == "ГП и ПФ":
                styler(row, 16,True,'FBF9EC', 3)
            elif row[0].value in mf_goodies_lvl2:
                styler(row, 16,True,'FBF9EC', 6)
            elif row[0].value in mf_goodies_lvl3:
                styler(row, 16, True, 'FBF9EC', 9)
            elif row[0].value in mf_goodies_lvl4:
                styler(row, 16, True, 'FBF9EC', 12)
            elif row[0].value in mf_goodies_lvl5:
                styler(row, 16, True, 'FBF9EC', 15)
            else:
                styler(row, 16,False,'FFFFFF', 18)
    wb.save(PATH + 'TestChanges2.xlsx')